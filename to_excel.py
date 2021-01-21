
import mysql.connector
from mysql.connector import errorcode
from xlrd import *
from xlsxwriter import *
import openpyxl


class Database:
    def export_to_xlsx(self):
        self.db = mysql.connector.Connect(
            host='localhost',
            port=1999,
            database='lazisnu',
            user='root',
            password='admin',
            auth_plugin='mysql_native_password')

        if self.db.is_connected:
            self.cur = self.db.cursor()
            # self.cur.execute('SELECT * FROM data ORDER BY dusun ,rw ,nama ASC')
            self.cur.execute(
                'SELECT * FROM data ORDER BY dusun, rw ASC, keterangan ASC;')

            rst = self.cur.fetchall()
            self.write_to_xlsx(rst)
            print('done')

    def write_to_xlsx(self, data):
        wb = Workbook('test1.xlsx')
        sheet1 = wb.add_worksheet()
        # merge header
        merge_format = wb.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
        })
        sheet1.merge_range('A1:N1', 'DATA LAZISNU', merge_format)
        cell_format = wb.add_format({
            'align': 'center',
        })
        # sub header
        sheet1.write(1, 0, 'NO', merge_format)
        sheet1.write(1, 1, 'KECAMATAN', merge_format)
        sheet1.write(1, 2, 'DESA', merge_format)
        sheet1.write(1, 3, 'DUSUN', merge_format)
        sheet1.write(1, 4, 'RT', merge_format)
        sheet1.write(1, 5, 'RW', merge_format)
        sheet1.write(1, 6, 'NIK', merge_format)
        sheet1.write(1, 7, 'TTL', merge_format)
        sheet1.write(1, 8, 'NPWP', merge_format)
        sheet1.write(1, 9, 'NO HP', merge_format)
        sheet1.write(1, 10, 'NAMA', merge_format)
        sheet1.write(1, 11, 'PEKERJAAN', merge_format)
        sheet1.write(1, 12, 'KETARANGAN', merge_format)
        sheet1.write(1, 13, 'STATUS', merge_format)

        row_number = 2
        for row in data:
            column_number = 0
            for item in row:
                if column_number in [0, 4, 5]:
                    sheet1.write(row_number, column_number,
                                 str(item), cell_format)
                else:
                    sheet1.write(row_number, column_number,
                                 str(item))
                column_number += 1
            row_number += 1
        wb.close()

    def read_xlsx(self):
        wb = openpyxl.load_workbook('test1.xlsx')
        ws = wb['MUNFIQ']
        dict_item = {}
        for row in range(3, ws.max_row + 1):
            tmp = []
            for cell in ws[f'A{row}:O{row}']:
                for x in cell:
                    tmp.append(x.value)
            print(tmp)
            dict_item[tmp[0]] = tmp

        return dict_item

    def store_to_db(self):
        data = self.read_xlsx()

        self.db = mysql.connector.Connect(
            host='localhost',
            port=1999,
            database='lazisnu',
            user='root',
            password='admin',
            auth_plugin='mysql_native_password')

        if self.db.is_connected:
            self.cur = self.db.cursor()

        tmp = []
        for idx, key in enumerate(data):
            tmp.append(data[key])

        query = 'INSERT INTO data VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
        self.cur.executemany(query, tmp)
        self.db.commit()
        print('Done')


db = Database()
db.export_to_xlsx()
# db.store_to_db()
